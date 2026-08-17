# -*- coding: utf-8 -*-
"""
excel-unidades.py - planilha completa das unidades A Formula.

SOMENTE LEITURA. Cruza:
  encontre-uma-loja_assets/lojas.json   -> o que esta no ar (cadastro)
  _coleta-google/{slug}.json            -> o que o Google publica (coleta 2026-08-17)

Saida: UNIDADES-A-FORMULA-{data}.xlsx com 5 abas.
Uso: python scripts/excel-unidades.py
"""
import io, json, os, re, unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
COLETA = os.path.join(ROOT, '_coleta-google')
DATA = '2026-08-17'
SITE = 'https://www.aformulabr.com.br'
DIAS = ['segunda-feira','terça-feira','quarta-feira','quinta-feira','sexta-feira','sábado','domingo']
# o que a pagina de unidade publica hoje (build-lojas.mjs -> HORARIO)
PUBLICADO = ['08:00-18:00']*5 + ['08:00-13:00', 'Fechado']


def ler_json(p):
    with io.open(p, encoding='utf-8') as f:
        return json.load(f)


def norm(s):
    s = unicodedata.normalize('NFD', str(s or ''))
    return ''.join(c for c in s if unicodedata.category(c) != 'Mn').lower()


def aberta(u):
    return 'em breve' not in norm(u.get('nome','') + ' ' + u.get('slug',''))


def rotulo(u):
    nome = (u.get('nome') or '').strip()
    cidade = (u.get('cidade') or '').strip()
    d = None
    if nome and norm(cidade) not in norm(nome):
        d = nome
    elif nome:
        partes = re.split(r'\s+[–—-]\s+', nome)
        d = ' — '.join(partes[1:]).replace('|', '—')
        d = re.sub(r'\s+', ' ', d).strip() or None
    return '%s%s' % (cidade, (' — ' + d) if d else '')


def numero(s):
    """Nº do logradouro. Tira o CEP ANTES: senao ele e confundido com o numero."""
    t = re.sub(r'\b\d{5}-?\d{3}\b', ' ', str(s or ''))
    t = re.sub(r'\bkm\s*\d+', ' ', t, flags=re.I)
    m = re.search(r',\s*(?:n[ºo]\.?\s*)?(\d{1,5})\s*[a-z]?\b', t, flags=re.I) or re.search(r'\b(\d{1,5})\s*[a-z]?\b', t)
    return m.group(1) if m else None


def so_num(s):
    return re.sub(r'\D', '', str(s or ''))


def legivel(h):
    return re.sub(r'(\d{2}:\d{2})(\d{2}:\d{2})', r'\1 e \2', str(h or ''))


# ---------- carga ----------
lojas = ler_json(os.path.join(ROOT, 'encontre-uma-loja_assets', 'lojas.json'))
col = {}
if os.path.isdir(COLETA):
    for f in os.listdir(COLETA):
        if f.endswith('.json'):
            j = ler_json(os.path.join(COLETA, f))
            col[j.get('slug')] = j

linhas = []
for u in lojas:
    slug = u.get('slug') or ''
    c = col.get(slug)
    op = aberta(u)
    graves, faltas = [], []
    if op:
        if not c or not c.get('confere'):
            graves.append('perfil do Google nao confirmado')
        elif c.get('diasCapturados') != 7:
            graves.append('perfil sem horario publicado')
        if not (u.get('telefone') or u.get('celular')):
            graves.append('sem telefone')
        if not u.get('cep'):
            faltas.append('CEP')
        if not u.get('email'):
            faltas.append('e-mail')
        if c and c.get('confere'):
            if not c.get('fotos'):
                faltas.append('sem foto no perfil')
            if not c.get('nota'):
                faltas.append('sem nota no Google')
            a, b = numero(c.get('enderecoGoogle')), numero(u.get('endereco'))
            if a and b and a != b:
                faltas.append('n. do endereco difere (cadastro %s x Google %s)' % (b, a))
    if not op:
        classe = 'Em breve'
    elif graves:
        classe = 'Critica'
    elif faltas:
        classe = 'Parcial'
    else:
        classe = 'Boa'

    tel_cad = u.get('celular') or u.get('telefone')
    tel_g = c.get('telefoneGoogle') if c else None
    if tel_g and tel_cad:
        tel_ok = 'igual' if so_num(tel_g) == so_num(tel_cad) else 'diferente'
    else:
        tel_ok = '-'
    a, b = (numero(c.get('enderecoGoogle')) if c else None), numero(u.get('endereco'))
    end_ok = ('igual' if a == b else 'diferente') if (a and b) else '-'

    linhas.append(dict(u=u, c=c, slug=slug, rot=rotulo(u), classe=classe,
                       pend='; '.join(graves + faltas), op=op,
                       tel_ok=tel_ok, end_ok=end_ok))

# ---------- estilos ----------
AZUL = 'FF063237'
CINZA = 'FFEEF2F2'
CORES = {'Boa': 'FFD9EFDD', 'Parcial': 'FFFFF3CD', 'Critica': 'FFF8D7DA', 'Em breve': 'FFE9ECEF'}
BORDA = Border(bottom=Side(style='thin', color='FFDDDDDD'))


def cabecalho(ws, cols):
    ws.append(cols)
    for i, _ in enumerate(cols, 1):
        cl = ws.cell(row=1, column=i)
        cl.font = Font(bold=True, color='FFFFFFFF', size=10)
        cl.fill = PatternFill('solid', fgColor=AZUL)
        cl.alignment = Alignment(vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'


def largura(ws, larguras):
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------- aba 1: Unidades ----------
ws = wb.active
ws.title = 'Unidades'
COLS = ['Slug', 'Unidade', 'Cidade', 'UF', 'Situacao', 'Classificacao', 'O que falta',
        'Endereco (cadastro)', 'CEP', 'Telefone', 'WhatsApp / celular', 'E-mail',
        'Latitude', 'Longitude', 'place_id (NAO confiavel)',
        'Nome no Google', 'Endereco no Google', 'Telefone no Google',
        'Nota', 'Avaliacoes', 'Fotos no perfil',
        'Telefone confere?', 'N. do endereco confere?',
        'Link do perfil no Google (fonte)', 'Pagina no site', 'Ficha no Second Brain', 'Coletado em']
cabecalho(ws, COLS)
for L in linhas:
    u, c = L['u'], L['c']
    ws.append([
        L['slug'], L['rot'], u.get('cidade'), u.get('estado'),
        'Em operacao' if L['op'] else 'Em breve', L['classe'], L['pend'] or '-',
        u.get('endereco'), u.get('cep'), u.get('telefone'), u.get('celular'), u.get('email'),
        u.get('lat'), u.get('lng'), u.get('place_id'),
        (c or {}).get('nomeGoogle'), (c or {}).get('enderecoGoogle'), (c or {}).get('telefoneGoogle'),
        float(c['nota']) if c and c.get('nota') else None,
        int(c['avaliacoes']) if c and c.get('avaliacoes') else None,
        len((c or {}).get('fotos') or []) if c else None,
        L['tel_ok'], L['end_ok'],
        (c or {}).get('urlPerfil'),
        ('%s/encontre-uma-loja/%s' % (SITE, L['slug'])) if L['op'] else '-',
        'TRABALHO/a-formula/unidades/%s.md' % L['slug'],
        (c or {}).get('coletadoEm') or '-',
    ])
largura(ws, [30, 34, 18, 6, 13, 14, 46, 52, 12, 18, 18, 34, 12, 12, 30,
             46, 52, 18, 7, 11, 8, 16, 20, 60, 52, 46, 12])
for r in range(2, ws.max_row + 1):
    classe = ws.cell(row=r, column=6).value
    fill = PatternFill('solid', fgColor=CORES.get(classe, 'FFFFFFFF'))
    for cx in range(1, len(COLS) + 1):
        cl = ws.cell(row=r, column=cx)
        cl.border = BORDA
        cl.alignment = Alignment(vertical='top', wrap_text=(cx in (7, 8, 16, 17)))
    ws.cell(row=r, column=6).fill = fill
    for cx in (24, 25):
        v = ws.cell(row=r, column=cx).value
        if v and str(v).startswith('http'):
            ws.cell(row=r, column=cx).hyperlink = v
            ws.cell(row=r, column=cx).font = Font(color='FF0563C1', underline='single', size=10)
ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(COLS)), ws.max_row)

# ---------- aba 2: Horarios ----------
ws2 = wb.create_sheet('Horarios')
C2 = ['Slug', 'Unidade', 'UF', 'Fonte'] + [d.capitalize() for d in DIAS] + ['Bate com o publicado?']
cabecalho(ws2, C2)
ws2.append(['-', 'PUBLICADO HOJE NO SITE (generico)', '-', 'build-lojas.mjs'] + PUBLICADO + ['-'])
for i in range(1, len(C2) + 1):
    cl = ws2.cell(row=2, column=i)
    cl.font = Font(bold=True, size=10)
    cl.fill = PatternFill('solid', fgColor='FFFFF3CD')
for L in linhas:
    c = L['c']
    if not (c and c.get('confere')):
        continue
    h = c.get('horarios') or {}
    if len(h) != 7:
        ws2.append([L['slug'], L['rot'], L['u'].get('estado'), 'perfil sem horario'] + ['-'] * 7 + ['-'])
        continue
    vals = [legivel(h.get(d)) for d in DIAS]
    bate = 'sim' if [v.replace(' ', '') for v in vals] == PUBLICADO else 'NAO'
    ws2.append([L['slug'], L['rot'], L['u'].get('estado'), 'Google ' + str(c.get('coletadoEm'))] + vals + [bate])
largura(ws2, [30, 34, 6, 20, 17, 17, 17, 17, 17, 17, 17, 20])
for r in range(3, ws2.max_row + 1):
    if ws2.cell(row=r, column=12).value == 'NAO':
        ws2.cell(row=r, column=12).fill = PatternFill('solid', fgColor='FFF8D7DA')
    for cx in range(1, len(C2) + 1):
        ws2.cell(row=r, column=cx).border = BORDA
ws2.auto_filter.ref = 'A1:L%d' % ws2.max_row

# ---------- aba 3: Pendencias ----------
ws3 = wb.create_sheet('Pendencias')
C3 = ['Unidade', 'UF', 'Classificacao', 'O que falta', 'Quem resolve']
cabecalho(ws3, C3)
for L in sorted([x for x in linhas if x['pend'] and x['op']],
                key=lambda x: (0 if x['classe'] == 'Critica' else 1, x['rot'])):
    quem = 'Conferir no Google (eu)' if 'nao confirmado' in L['pend'] else (
        'Unidade / franqueado' if 'sem horario' in L['pend'] else 'Matriz (cadastro)')
    ws3.append([L['rot'], L['u'].get('estado'), L['classe'], L['pend'], quem])
largura(ws3, [34, 6, 14, 60, 26])
for r in range(2, ws3.max_row + 1):
    ws3.cell(row=r, column=3).fill = PatternFill('solid', fgColor=CORES.get(ws3.cell(row=r, column=3).value, 'FFFFFFFF'))
    for cx in range(1, len(C3) + 1):
        ws3.cell(row=r, column=cx).border = BORDA
        ws3.cell(row=r, column=cx).alignment = Alignment(vertical='top', wrap_text=(cx == 4))
ws3.auto_filter.ref = 'A1:E%d' % ws3.max_row

# ---------- aba 4: Fotos ----------
ws4 = wb.create_sheet('Fotos')
C4 = ['Unidade', 'UF', 'Qtd', 'URL da foto (perfil do Google)']
cabecalho(ws4, C4)
for L in linhas:
    c = L['c']
    if not (c and c.get('confere')):
        continue
    for url in (c.get('fotos') or []):
        ws4.append([L['rot'], L['u'].get('estado'), len(c.get('fotos') or []), url])
largura(ws4, [34, 6, 6, 120])
for r in range(2, ws4.max_row + 1):
    cl = ws4.cell(row=r, column=4)
    if cl.value:
        cl.hyperlink = cl.value
        cl.font = Font(color='FF0563C1', underline='single', size=9)
    for cx in range(1, len(C4) + 1):
        ws4.cell(row=r, column=cx).border = BORDA
ws4.auto_filter.ref = 'A1:D%d' % ws4.max_row

# ---------- aba 5: Leia-me ----------
ws5 = wb.create_sheet('Leia-me')
ws5.column_dimensions['A'].width = 120
tot = len([x for x in linhas if x['op']])
cont = {k: len([x for x in linhas if x['classe'] == k]) for k in ('Boa', 'Parcial', 'Critica', 'Em breve')}
texto = [
    ('t', 'Unidades A Fórmula — dado do cadastro × dado do Google'),
    ('', 'Gerado em %s. Documento de LEITURA: nada foi alterado no cadastro, nas páginas ou no ar.' % DATA),
    ('', ''),
    ('h', 'Fontes'),
    ('', '• Cadastro ("o que está no ar"): encontre-uma-loja_assets/lojas.json — é o arquivo que alimenta o mapa'),
    ('', '  do /encontre-uma-loja e os cards com telefone e endereço.'),
    ('', '• Google: perfil público de cada unidade no Google Maps, coletado em %s.' % DATA),
    ('', '  A coluna "Link do perfil no Google" é a fonte exata de onde cada dado saiu — dá pra conferir clicando.'),
    ('', ''),
    ('h', 'Onde o mapeamento está'),
    ('', '• %d unidades em operação + %d "em breve" = %d no cadastro.' % (tot, cont['Em breve'], len(linhas))),
    ('', '• Boas (nada faltando): %d' % cont['Boa']),
    ('', '• Parciais (falta detalhe, dá pra publicar): %d' % cont['Parcial']),
    ('', '• Críticas (não publicar sem resolver): %d' % cont['Critica']),
    ('', ''),
    ('h', 'Três avisos que mudam como ler esta planilha'),
    ('', '1) O place_id do cadastro NÃO é confiável. Em salvador-shopping-paralela ele aponta para o PRÉDIO'),
    ('', '   (categoria "Edifício"), não para a loja — colher por ele traz a nota do Shopping Paralela'),
    ('', '   (4,5 / 35.912 avaliações) no lugar da farmácia (4,2). Por isso a coleta entrou por busca de nome.'),
    ('', '2) Telefone repetido entre unidades é ESPERADO: lojas do mesmo grupo usam o mesmo call center.'),
    ('', '   Quando a coluna "Telefone confere?" diz "diferente", não é erro — o cadastro traz o call center'),
    ('', '   e o Google traz o fixo da loja. São canais diferentes; qual publicar é decisão do operador.'),
    ('', '3) O horário publicado hoje no site é GENÉRICO (08:00–18:00, sáb 08:00–13:00, dom fechado) e bate'),
    ('', '   com apenas 4 das 56 unidades que têm semana completa no Google. Ver aba "Horarios".'),
    ('', ''),
    ('h', 'Limite da coleta'),
    ('', '• A semana completa só sai do Google em sessão logada e janela real de navegador; headless recebe'),
    ('', '  "visualização limitada" com 1 dia só, e google.com/search devolve CAPTCHA.'),
    ('', '• 14 unidades não têm horário publicado no perfil — só o franqueado informa.'),
    ('', '• 5 unidades não foram confirmadas: a busca não chegou num perfil cujo nome comece com "A Fórmula",'),
    ('', '  então nada foi aproveitado. A trava é estrita de propósito — a versão frouxa aceitaria concorrente'),
    ('', '  ("Companhia da Fórmula Alecrim — Farmácia de Manipulação", que apareceu como anúncio patrocinado).'),
    ('', '• Fotos: só as URLs foram registradas. Nenhuma imagem foi baixada nem usada.'),
    ('', '• Nota e nº de avaliações mudam sozinhos; valem para o dia da coleta.'),
]
for tipo, linha in texto:
    ws5.append([linha])
    cl = ws5.cell(row=ws5.max_row, column=1)
    if tipo == 't':
        cl.font = Font(bold=True, size=15, color=AZUL)
    elif tipo == 'h':
        cl.font = Font(bold=True, size=11, color=AZUL)
    else:
        cl.font = Font(size=10)
    cl.alignment = Alignment(vertical='top')

saida = os.path.join(ROOT, 'UNIDADES-A-FORMULA-%s.xlsx' % DATA)
wb.save(saida)
print('escrito:', saida)
print('Unidades: %d linhas | Horarios: %d | Pendencias: %d | Fotos: %d'
      % (ws.max_row - 1, ws2.max_row - 2, ws3.max_row - 1, ws4.max_row - 1))
print('Boa %d | Parcial %d | Critica %d | Em breve %d'
      % (cont['Boa'], cont['Parcial'], cont['Critica'], cont['Em breve']))
